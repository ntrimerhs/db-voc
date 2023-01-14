import sys
from openpyxl import Workbook, load_workbook

wb = load_workbook('mega.xlsx') 
ws = wb.active

ws['C5'] = sys.argv[1]
ws['D5'] = sys.argv[2]
ws['E5'] = sys.argv[3]
ws['G5'] = sys.argv[10]
ws['H5'] = sys.argv[12]

wb.save('mega.xlsx')
sys.stdout.flush()