import sys
from openpyxl import Workbook, load_workbook

wb = load_workbook('./../excel/pritsoulhs1/pritsoulhs2.xlsx') 
ws = wb.active

kales = int(sys.argv[1])
metries = int(sys.argv[2])
kakes = int(sys.argv[3])
epistrofes = int(sys.argv[10])
sxolia = int(sys.argv[12])

print(type(kales), type(metries), type(kakes), type(epistrofes), type(sxolia))
print(kales, metries, kakes, epistrofes, sxolia)

ws['C5'] = kales 
ws['D5'] = metries
ws['E5'] = kakes 
ws['G5'] = epistrofes
ws['H5'] = sxolia

wb.save('./../excel/pritsoulhs1/pritsoulhs2.xlsx') 
sys.stdout.flush()